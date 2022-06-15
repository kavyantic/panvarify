
from cmath import log
from xml import dom
from xml.dom.minidom import Document
from numpy import var
import selenium
from selenium.webdriver import Chrome,ChromeOptions
from time import sleep
from selenium.webdriver.common import keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import os
import pickle
import selenium.common.exceptions as exce

xlsxFile = ""
baseUrl = r"https://www.tdscpc.gov.in/app/ded/panverify.xhtml"
options = ChromeOptions()
options.add_argument = {'user-data-dir':r'C:\Users\kavya\AppData\Local\Google\Chrome\User Data\Default'}
driver = Chrome("./chromedriver.exe",options=options)

def login():
	driver.get('https://www.tdscpc.gov.in/app/login.xhtml')
	userId = driver.find_element(By.XPATH,'//*[@id="userId"]')
	password = driver.find_element(By.XPATH,'//*[@id="psw"]')
	tanId = driver.find_element(By.XPATH,'//*[@id="tanpan"]')
	loginBtn = driver.find_element(By.XPATH,'//*[@id="clickLogin"]')
	userId.send_keys('LKNE05333G1')
	password.send_keys('A9839573080')
	tanId.send_keys('LKNE05333G')
	loginBtn.click()
	wait = WebDriverWait(driver,300).until(lambda driver:driver.current_url != "https://www.tdscpc.gov.in/app/login.xhtml")
	pickle.dump(driver.get_cookies(),open("cookies.pkl","wb"))

for file in os.listdir():
	if(file.split(".")[-1]=="xlsx"):
		xlsxFile = os.getcwd() + "\\" +file

driver.get(baseUrl)

try:
	cookies = pickle.load(open("cookies.pkl", "rb"))
	driver.delete_all_cookies()
	for cookie in cookies: 
		driver.add_cookie(cookie)
	driver.get(baseUrl)
except FileNotFoundError:
	print("file not found")
	pass

if(driver.current_url == "https://www.tdscpc.gov.in/app/login.xhtml"):
	login()

read_excel = openpyxl.load_workbook(xlsxFile) 
read_sheet = read_excel.active
write_excel = openpyxl.load_workbook(filename=xlsxFile)
write_sheet = write_excel['Sheet1'] 

def updateExcel(field):
	write_sheet.cell(row=field['row'],column=8,value=field['match'])
	write_sheet.cell(row=field['row'],column=9,value=field['mismatch'])
	write_sheet.cell(row=field['row'],column=10,value=field['invalid'])
	write_sheet.cell(row=field['row'],column=7,value=field['name'])


	try:
		write_excel.save(xlsxFile)
	except PermissionError: 
		input(f"Please exit the excel file then press enter :")
		write_excel.save(xlsxFile)


write_sheet.cell(row=1,column=12,value="1")
for idx,row in enumerate(read_sheet.iter_rows(),1):
	pan_no = row[4].value
	name = row[5].value
	print(name)
	verified = bool(row[9].value)
	if(verified):
		continue

	panInput = driver.find_element(By.XPATH,'//*[@id="pannumber"]')
	formType = driver.find_element(By.XPATH,'//*[@id="frmType1"]')
	goButton = driver.find_element(By.XPATH,'//*[@id="clickGo1"]')
	fields = {"row":idx,"name":"","match":"false","mismatch":"false","invalid":"false"}


	
	panInput.clear()
	panInput.send_keys(pan_no)
	formType.send_keys('24Q')
	goButton.click()
	try:
		wait  = WebDriverWait(driver,2).until(EC.visibility_of_all_elements_located((By.XPATH,'//*[@id="name"]')))
		web_name = driver.find_element(By.XPATH,'//*[@id="name"]').text
	except selenium.common.exceptions.TimeoutException:
		fields["invalid"] = "true"
		updateExcel(fields)
		continue
	 
	
	if(name and name.lower()==web_name.lower()):
		print(name," ",web_name)
		fields["match"] = "true"
		fields["name"] = web_name
		updateExcel(fields)
	else:
		fields["mismatch"] = "true"	
		fields["name"] = web_name
		updateExcel(fields)

